import numpy as np
from openpyxl import load_workbook
from IPM_MVIP.coefficient_optimiser2 import run

wb = load_workbook('../../data.xlsx')
ws = wb.get_sheet_by_name('SUMMARY')

# upper derivative
# rates for dx6. rates = R7
# R7 = x6
R7 = []
itera = ws.iter_rows('S5:S103')
for row in itera:
    for cell in row:
        R7.append(float(cell.value))

rates1 = np.array([R7]).T  # Take a transpose. To make each row is a data entry for one time step.


# der1 = dx6
der1 = []
itera = ws.iter_rows('Z5:Z103')
for row in itera:
    for cell in row:
        der1.append(float(cell.value))
der1 = np.array(der1)


# div1 = x6
# x6
x6 = []
itera = ws.iter_rows('L5:L103')
for row in itera:
    for cell in row:
        x6.append(float(cell.value))
div1 = np.array([x6]).T


# lower derivative
# rates for dx4. rates = R4
R4 = []
itera = ws.iter_rows('P5:P103')
for row in itera:
    for cell in row:
        R4.append(float(cell.value))

rates2 = np.array([R4]).T  # Take a transpose. To make each row is a data entry for one time step.

# der2 = dx4
der2 = []
itera = ws.iter_rows('X5:X103')
for row in itera:
    for cell in row:
        der2.append(float(cell.value))
der2 = np.array(der2)

# div2 = x4
# x4
x4 = []
itera = ws.iter_rows('J5:J103')
for row in itera:
    for cell in row:
        x4.append(float(cell.value))
div2 = np.array([x4]).T

run(1000, rates2, rates1, der2, der1, div2, div1)
# run(1000, rates1, rates2, der1, der2, div1, div2)
