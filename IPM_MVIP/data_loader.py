from openpyxl import load_workbook

import numpy as np


def load_rates(work_sheet, locations):

    # work_sheet is returned by openpyxl.load_worksheet.
    # Information about openpyxl can be found on:
    # https://openpyxl.readthedocs.org/en/default/tutorial.html#loading-from-a-file

    # locations is a list strings, each string specify the location of data for one variable/rate/derivative.
    # the data for one variable needs to be placed in a column without spacing.
    # For example, ['G5:G105', 'Z5:Z105'] specify locations of two variables. first variable's data is located in cells from G5 to G105

    rates = []
    for loc in locations:
        rate = []
        itera = work_sheet.iter_rows(loc)
        for row in itera:
            for cell in row:
                rate.append(float(cell.value))
        rates.append(np.array(rate))

    return np.array(rates).T  # transpose the 2 dimensional array to make each row as a data for time step.


def load_single_variable(work_sheet, location, all_one=False):
    # location is one string.
    # all_one: if it is true, this method will return all

    der = []
    itera = work_sheet.iter_rows(location)
    for row in itera:
        for cell in row:
            if all_one:
                der.append(1)
            else:
                der.append(float(cell.value))
    der = np.array(der)

    return der


def load_o_fragment(work_sheet, locations):
    divs = []
    for loc in locations:
        itera = work_sheet.iter_rows(loc)
        div = []
        for row in itera:
            for cell in row:
                div.append(float(cell.value))
        divs.append(np.array(div))
    return np.array(divs).T  # transpose


def load_equation_rates_data(worksheet, equation, locations):

    # der: dx1
    der1 = load_single_variable(worksheet, locations['d' + equation.dependent_variable])

    # rates
    rates_data = []
    for rate in equation.rates:
        var_datas = []
        for variable in rate.variables:
            var_datas.append(load_single_variable(worksheet, locations[variable]))

        rate_data = np.array([1]*len(var_datas[0]))
        for var_data in var_datas:
            rate_data = rate_data * var_data

        rates_data.append(rate_data)

    return np.array(rates_data).T
